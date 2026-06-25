"""
reporter.py
Gera relatório .xlsx após cada lote processado.

Abas:
  1. Lote        — uma linha por PDF combinado gerado
  2. Vencimentos — calendário de pagamentos ordenado por data
  3. Resumo      — totais por fornecedor
  4. Graficos    — dados para gráficos (Power BI conecta direto)
"""

from pathlib import Path
from datetime import datetime, date
from typing import Optional
import re

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import DataPoint
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ---------------------------------------------------------------------------
# Helpers de formatação
# ---------------------------------------------------------------------------

NAV_BG   = "1E2433"
AZUL     = "2563EB"
AZUL_CLR = "DBEAFE"
CINZA    = "F0F2F5"
BORDA    = "E5E7EB"
VERDE    = "10B981"
VERDE_CLR= "ECFDF5"
AMBER    = "F59E0B"
AMBER_CLR= "FFFBEB"
BRANCO   = "FFFFFF"
TEXTO    = "111827"
SUBTEXTO = "6B7280"


def _borda_fina():
    lado = Side(style="thin", color=BORDA)
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _cabecalho(ws, colunas: list, linha: int = 1, cor_bg: str = NAV_BG, cor_txt: str = BRANCO):
    for col, titulo in enumerate(colunas, 1):
        c = ws.cell(row=linha, column=col, value=titulo)
        c.fill = PatternFill("solid", fgColor=cor_bg)
        c.font = Font(bold=True, color=cor_txt, size=10, name="Segoe UI")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _borda_fina()


def _celula(ws, linha, col, valor, negrito=False, cor_bg=None, alinhamento="left", formato=None):
    c = ws.cell(row=linha, column=col, value=valor)
    c.font = Font(bold=negrito, size=10, name="Segoe UI", color=TEXTO)
    c.alignment = Alignment(horizontal=alinhamento, vertical="center")
    c.border = _borda_fina()
    if cor_bg:
        c.fill = PatternFill("solid", fgColor=cor_bg)
    if formato:
        c.number_format = formato
    return c


def _parse_data(data_str: Optional[str]) -> Optional[date]:
    if not data_str:
        return None
    m = re.match(r"(\d{2})/(\d{2})/(\d{4})", data_str)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            return None
    return None


def _larguras(ws, larguras: dict):
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 32


# ---------------------------------------------------------------------------
# Geração do relatório
# ---------------------------------------------------------------------------

def gerar_relatorio(resultado_matching, resultados_merger: list, pasta_saida: str) -> Optional[str]:
    """
    Gera o arquivo .xlsx na pasta_saida.
    Retorna o caminho do arquivo gerado ou None em caso de erro.
    """
    if not OPENPYXL_OK:
        return None

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove aba padrão

    _aba_lote(wb, resultado_matching, resultados_merger)
    _aba_vencimentos(wb, resultado_matching, resultados_merger)
    _aba_resumo(wb, resultado_matching, resultados_merger)
    _aba_graficos(wb, resultado_matching, resultados_merger)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome = f"relatorio_lote_{ts}.xlsx"
    Path(pasta_saida).mkdir(parents=True, exist_ok=True)
    caminho = str(Path(pasta_saida) / nome)

    wb.save(caminho)
    return caminho


# ---------------------------------------------------------------------------
# Aba 1: Lote
# ---------------------------------------------------------------------------

def _aba_lote(wb, resultado, resultados_merger):
    ws = wb.create_sheet("Lote")
    ws.freeze_panes = "A2"

    cols = [
        "Arquivo Gerado", "Status", "NF Nº", "Emitente", "Descrição Produto",
        "NCM", "Quantidade (t)", "Valor NF (R$)", "Valor Unit. (R$)",
        "Vencimento NF", "Nº Pedido SAP", "Vencimento Boleto", "Valor Boleto (R$)",
        "Tipo Combinação", "Alertas"
    ]
    _cabecalho(ws, cols)

    for res in resultados_merger:
        par = res["par"]
        nf  = par.nf
        sap = par.sap
        bol = par.boleto
        ok  = res["sucesso"]
        lin = ws.max_row + 1
        cor = VERDE_CLR if ok else AMBER_CLR

        dados = [
            res.get("nome_arquivo") or res.get("erro", ""),
            "Gerado" if ok else "Erro",
            nf.numero_nf,
            nf.emitente,
            nf.descricao_produto,
            nf.ncm_produto,
            nf.quantidade_ton,
            nf.valor_nf,
            nf.valor_unitario,
            _parse_data(nf.vencimento_nf),
            sap.numero_pedido if sap else None,
            _parse_data(bol.vencimento_boleto) if bol else None,
            bol.valor_boleto if bol else None,
            par.tipo,
            " | ".join(par.alertas) if par.alertas else "",
        ]

        formatos = [
            None, None, None, None, None, None,
            '#,##0.000 "t"', 'R$ #,##0.00', 'R$ #,##0.00',
            'DD/MM/AAAA', None, 'DD/MM/AAAA', 'R$ #,##0.00',
            None, None
        ]

        for col, (val, fmt) in enumerate(zip(dados, formatos), 1):
            _celula(ws, lin, col, val, cor_bg=cor, formato=fmt,
                    alinhamento="center" if col in (1, 2, 3, 6, 10, 11, 12, 14) else "left")

    _larguras(ws, {
        "A": 38, "B": 10, "C": 8, "D": 36, "E": 32,
        "F": 12, "G": 14, "H": 16, "I": 16, "J": 14,
        "K": 16, "L": 16, "M": 16, "N": 12, "O": 40
    })


# ---------------------------------------------------------------------------
# Aba 2: Vencimentos
# ---------------------------------------------------------------------------

def _aba_vencimentos(wb, resultado, resultados_merger):
    ws = wb.create_sheet("Vencimentos")
    ws.freeze_panes = "A2"

    cols = ["Data Vencimento", "Dias para Vencer", "Nº Pedido", "NF Nº",
            "Emitente", "Descrição Produto", "Valor NF (R$)", "Valor Boleto (R$)", "Status"]
    _cabecalho(ws, cols)

    hoje = date.today()
    linhas = []

    for res in resultados_merger:
        if not res["sucesso"]:
            continue
        par = res["par"]
        nf  = par.nf
        bol = par.boleto
        sap = par.sap

        dt = _parse_data(nf.vencimento_nf)
        if not dt:
            continue

        dias = (dt - hoje).days
        if dias < 0:
            status = "Vencido"
            cor = "FEE2E2"
        elif dias <= 3:
            status = "Vence em breve"
            cor = AMBER_CLR
        else:
            status = "A vencer"
            cor = VERDE_CLR

        linhas.append((dt, dias, sap.numero_pedido if sap else None,
                       nf.numero_nf, nf.emitente, nf.descricao_produto,
                       nf.valor_nf, bol.valor_boleto if bol else None, status, cor))

    # Ordena por data
    linhas.sort(key=lambda x: x[0])

    fmts = ['DD/MM/AAAA', '0 "dias"', None, None, None, None, 'R$ #,##0.00', 'R$ #,##0.00', None]
    aligns = ["center", "center", "center", "center", "left", "left", "right", "right", "center"]

    for dt, dias, pedido, nf_num, emit, desc, val_nf, val_bol, status, cor in linhas:
        lin = ws.max_row + 1
        vals = [dt, dias, pedido, nf_num, emit, desc, val_nf, val_bol, status]
        for col, (v, fmt, aln) in enumerate(zip(vals, fmts, aligns), 1):
            _celula(ws, lin, col, v, cor_bg=cor, formato=fmt, alinhamento=aln)

    _larguras(ws, {
        "A": 16, "B": 16, "C": 16, "D": 8,
        "E": 36, "F": 32, "G": 18, "H": 18, "I": 16
    })


# ---------------------------------------------------------------------------
# Aba 3: Resumo por fornecedor
# ---------------------------------------------------------------------------

def _aba_resumo(wb, resultado, resultados_merger):
    ws = wb.create_sheet("Resumo")
    ws.freeze_panes = "A2"

    # Agrega por emitente
    por_emit = {}
    for res in resultados_merger:
        if not res["sucesso"]:
            continue
        nf = res["par"].nf
        key = nf.emitente or "Desconhecido"
        if key not in por_emit:
            por_emit[key] = {
                "qtd_nfs": 0, "total_ton": 0.0, "total_valor": 0.0,
                "produtos": set(), "pedidos": set()
            }
        por_emit[key]["qtd_nfs"] += 1
        por_emit[key]["total_ton"] += nf.quantidade_ton or 0
        por_emit[key]["total_valor"] += nf.valor_nf or 0
        if nf.descricao_produto:
            por_emit[key]["produtos"].add(nf.descricao_produto)
        if res["par"].sap and res["par"].sap.numero_pedido:
            por_emit[key]["pedidos"].add(res["par"].sap.numero_pedido)

    cols = ["Fornecedor", "Qtd NFs", "Total Ton.", "Total Valor (R$)",
            "Ticket Médio (R$)", "Produtos", "Nº Pedidos"]
    _cabecalho(ws, cols)

    cores = [CINZA, BRANCO]
    for i, (emit, d) in enumerate(sorted(por_emit.items(), key=lambda x: -x[1]["total_valor"])):
        lin = ws.max_row + 1
        cor = cores[i % 2]
        ticket = d["total_valor"] / d["qtd_nfs"] if d["qtd_nfs"] else 0
        vals = [
            emit, d["qtd_nfs"], d["total_ton"], d["total_valor"],
            ticket, ", ".join(sorted(d["produtos"])), len(d["pedidos"])
        ]
        fmts = [None, None, '#,##0.000 "t"', 'R$ #,##0.00', 'R$ #,##0.00', None, None]
        aligns = ["left", "center", "center", "right", "right", "left", "center"]
        for col, (v, fmt, aln) in enumerate(zip(vals, fmts, aligns), 1):
            _celula(ws, lin, col, v, cor_bg=cor, formato=fmt, alinhamento=aln)

    # Linha de totais
    lin = ws.max_row + 1
    total_ton = sum(d["total_ton"] for d in por_emit.values())
    total_val = sum(d["total_valor"] for d in por_emit.values())
    total_nfs = sum(d["qtd_nfs"] for d in por_emit.values())
    for col, (v, fmt, aln) in enumerate(zip(
        ["TOTAL", total_nfs, total_ton, total_val, "", "", ""],
        [None, None, '#,##0.000 "t"', 'R$ #,##0.00', None, None, None],
        ["left", "center", "center", "right", "right", "left", "center"]
    ), 1):
        c = _celula(ws, lin, col, v, negrito=True, cor_bg=NAV_BG, formato=fmt, alinhamento=aln)
        c.font = Font(bold=True, color=BRANCO, size=10, name="Segoe UI")

    _larguras(ws, {"A": 42, "B": 10, "C": 14, "D": 18, "E": 18, "F": 40, "G": 12})


# ---------------------------------------------------------------------------
# Aba 4: Gráficos (dados estruturados para Power BI / gráfico embutido)
# ---------------------------------------------------------------------------

def _aba_graficos(wb, resultado, resultados_merger):
    ws = wb.create_sheet("Graficos")

    sucessos = [r for r in resultados_merger if r["sucesso"]]

    # --- Seção 1: Valores por NF ---
    ws["A1"] = "VALORES POR NF"
    ws["A1"].font = Font(bold=True, size=11, color=NAV_BG, name="Segoe UI")
    _cabecalho(ws, ["NF Nº", "Emitente", "Valor NF (R$)", "Quantidade (t)"], linha=2)

    for res in sucessos:
        nf = res["par"].nf
        lin = ws.max_row + 1
        cor = CINZA if lin % 2 == 0 else BRANCO
        for col, (v, fmt) in enumerate(zip(
            [nf.numero_nf, nf.emitente, nf.valor_nf, nf.quantidade_ton],
            [None, None, 'R$ #,##0.00', '#,##0.000 "t"']
        ), 1):
            _celula(ws, lin, col, v, cor_bg=cor, formato=fmt,
                    alinhamento="center" if col in (1, 4) else "left")

    # Gráfico de barras — Valor por NF
    if len(sucessos) >= 2:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Valor por NF (R$)"
        chart.style = 10
        chart.y_axis.title = "Valor (R$)"
        chart.x_axis.title = "NF"
        chart.height = 12
        chart.width = 22

        n = len(sucessos)
        data_ref   = Reference(ws, min_col=3, min_row=2, max_row=2 + n)
        labels_ref = Reference(ws, min_col=1, min_row=3, max_row=2 + n)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(labels_ref)
        chart.series[0].graphicalProperties.solidFill = AZUL
        ws.add_chart(chart, "F2")

    # --- Seção 2: TOP 5 Produtos ---
    linha_ini = ws.max_row + 3
    ws.cell(row=linha_ini, column=1, value="TOP 5 PRODUTOS").font = Font(
        bold=True, size=11, color=NAV_BG, name="Segoe UI")
    _cabecalho(ws, ["Produto", "Qtd NFs", "Total Ton.", "Total Valor (R$)"], linha=linha_ini + 1)

    por_prod = {}
    for res in sucessos:
        nf = res["par"].nf
        key = nf.descricao_produto or "N/D"
        if key not in por_prod:
            por_prod[key] = {"qtd": 0, "ton": 0.0, "valor": 0.0}
        por_prod[key]["qtd"] += 1
        por_prod[key]["ton"] += nf.quantidade_ton or 0
        por_prod[key]["valor"] += nf.valor_nf or 0

    top5 = sorted(por_prod.items(), key=lambda x: -x[1]["valor"])[:5]
    linha_top = linha_ini + 2
    for i, (prod, d) in enumerate(top5):
        cor = CINZA if i % 2 == 0 else BRANCO
        for col, (v, fmt) in enumerate(zip(
            [prod, d["qtd"], d["ton"], d["valor"]],
            [None, None, '#,##0.000 "t"', 'R$ #,##0.00']
        ), 1):
            _celula(ws, linha_top + i, col, v, cor_bg=cor, formato=fmt,
                    alinhamento="left" if col == 1 else "center")

    # Gráfico TOP 5 Produtos
    if top5:
        chart2 = BarChart()
        chart2.type = "bar"
        chart2.title = "Top 5 Produtos - Valor Total (R$)"
        chart2.style = 10
        chart2.height = 10
        chart2.width = 22

        n2 = len(top5)
        data2   = Reference(ws, min_col=4, min_row=linha_ini + 1, max_row=linha_ini + 1 + n2)
        labels2 = Reference(ws, min_col=1, min_row=linha_ini + 2, max_row=linha_ini + 1 + n2)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(labels2)
        chart2.series[0].graphicalProperties.solidFill = "10B981"
        ws.add_chart(chart2, f"F{linha_ini}")

    # --- Seção 3: Resumo por fornecedor (para Power BI) ---
    linha_forn = ws.max_row + 3
    ws.cell(row=linha_forn, column=1, value="RESUMO POR FORNECEDOR").font = Font(
        bold=True, size=11, color=NAV_BG, name="Segoe UI")
    _cabecalho(ws, ["Fornecedor", "Qtd NFs", "Total Valor (R$)", "% do Total"],
               linha=linha_forn + 1)

    por_emit = {}
    total_geral = sum((r["par"].nf.valor_nf or 0) for r in sucessos)
    for res in sucessos:
        nf = res["par"].nf
        key = nf.emitente or "Desconhecido"
        por_emit.setdefault(key, {"qtd": 0, "valor": 0.0})
        por_emit[key]["qtd"] += 1
        por_emit[key]["valor"] += nf.valor_nf or 0

    for i, (emit, d) in enumerate(sorted(por_emit.items(), key=lambda x: -x[1]["valor"])):
        cor = CINZA if i % 2 == 0 else BRANCO
        pct = d["valor"] / total_geral if total_geral else 0
        lin = linha_forn + 2 + i
        for col, (v, fmt) in enumerate(zip(
            [emit, d["qtd"], d["valor"], pct],
            [None, None, 'R$ #,##0.00', '0.0%']
        ), 1):
            _celula(ws, lin, col, v, cor_bg=cor, formato=fmt,
                    alinhamento="left" if col == 1 else "center")

    _larguras(ws, {"A": 38, "B": 10, "C": 18, "D": 14, "E": 4,
                   "F": 6, "G": 6, "H": 6, "I": 6, "J": 6,
                   "K": 6, "L": 6, "M": 6, "N": 6})